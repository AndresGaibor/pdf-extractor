import pymupdf4llm
import re
import tempfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from constants import NUMERACION
from extractors import obtener_participante, extraer_cargo, extraer_organos_judiciales, extraer_provincias_explicitas, separar_tipo_y_localidad
from territorios import extraer_provincia_y_ccaa

# Para el archivo de excel
COLUMNAS = [
    "Participante",
    "Tipo de funcionaria",
    "Tipo de tribunal de origen",
    "Tipo de tribunal de destino",
    "Provincia/Localidad de origen",
    "Provincia/Localidad de destino",
]


def _resolver_localidad(localidad_texto, texto_tribunal_completo):
    """
    Dado un nombre de localidad extraído del tribunal, devuelve una cadena
    formateada como 'Localidad (Provincia)' o 'Provincia (CCAA)' según corresponda.

    - Si la localidad es un municipio → 'Municipio (Provincia)'
    - Si la localidad es una provincia → 'Provincia (CCAA)'
    - Si la localidad es una CCAA → 'CCAA'
    """
    if not localidad_texto:
        # Sin localidad separada, intentar extraer del texto completo del tribunal
        prov, ccaa = extraer_provincia_y_ccaa(texto_tribunal_completo)
        if prov and ccaa:
            return f"{prov} ({ccaa})"
        return prov or ccaa or ""

    prov, ccaa = extraer_provincia_y_ccaa(localidad_texto)

    # Si extraer_provincia_y_ccaa devuelve la provincia (=localidad es municipio)
    if prov and prov != localidad_texto:
        return f"{localidad_texto} ({prov})"
    # Si la localidad ES la provincia, mostrar la CCAA
    if prov and ccaa:
        return f"{localidad_texto} ({ccaa})"
    # Si solo devolvió CCAA (la localidad era una CCAA misma, como "Cataluña")
    if ccaa and not prov:
        return localidad_texto
    # Fallback
    return localidad_texto


def procesar_pdf(ruta_pdf: str) -> list[dict]:
    """
    Procesa un PDF y devuelve una lista de diccionarios con los datos extraídos.
    Acepta una ruta de archivo.
    """
    md_text = pymupdf4llm.to_markdown(ruta_pdf)
    return _extraer_datos(md_text)


def procesar_pdf_bytes(pdf_bytes: bytes, nombre_archivo: str = "temp.pdf") -> list[dict]:
    """
    Procesa bytes de un PDF y devuelve una lista de diccionarios con los datos extraídos.
    Útil para archivos subidos vía Streamlit.
    """
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name

    try:
        md_text = pymupdf4llm.to_markdown(tmp_path)
        return _extraer_datos(md_text)
    finally:
        os.unlink(tmp_path)


def _limpiar_cabeceras_boe(texto: str) -> str:
    """Elimina cabeceras y pies de página del BOE que rompen párrafos entre páginas."""
    # 1. Eliminar el bloque completo de cabecera del BOE (título + fecha/página)
    #    Ejemplo: # **BOLETÍN OFICIAL DEL ESTADO**\n\n**Núm. 35** ... **Sec. II.A.  Pág. 20186**
    texto = re.sub(
        r'#+\s*\*{0,2}\s*BOLETÍN\s+OFICIAL\s+DEL\s+ESTADO\s*\*{0,2}\s*'
        r'.*?Pág\.\s*\d+\s*\*{0,2}',
        ' ', texto, flags=re.DOTALL
    )
    # 2. Si quedó solo la línea de título sin la de página
    texto = re.sub(
        r'#+\s*\*{0,2}\s*BOLETÍN\s+OFICIAL\s+DEL\s+ESTADO\s*\*{0,2}',
        ' ', texto
    )
    # 3. Eliminar pie de página del BOE (url + nombre + ISSN)
    texto = re.sub(
        r'\*{0,2}\s*https?://www\.boe\.es\s*\*{0,2}.*?ISSN[:\s]*[\d\-X]+\s*\*{0,2}',
        ' ', texto, flags=re.DOTALL
    )
    # 4. Colapsar saltos de línea múltiples que quedaron tras la limpieza
    texto = re.sub(r'\n\s*\n\s*\n+', '\n\n', texto)
    return texto


def _extraer_datos(md_text: str) -> list[dict]:
    """Lógica central de extracción a partir del texto markdown del PDF."""
    # Limpiar cabeceras/pies de página del BOE que rompen párrafos
    md_text = _limpiar_cabeceras_boe(md_text)

    parrafos = [p.strip() for p in md_text.split('\n\n') if p.strip()]

    # Reconstruir párrafos que fueron cortados por saltos de página:
    # si un párrafo NO empieza con NUMERACION y el anterior SÍ, es una continuación
    parrafos_filtrados = []
    for parrafo in parrafos:
        if parrafo.startswith(tuple(NUMERACION)):
            parrafos_filtrados.append(parrafo)
        elif parrafos_filtrados:
            # Es un fragmento de continuación → unir al párrafo anterior
            parrafos_filtrados[-1] += ' ' + parrafo

    filas = []

    for parrafo in parrafos_filtrados:
        try:
            participante = obtener_participante(parrafo)

            if len(participante.split()) < 2:
                continue

            cargo = extraer_cargo(parrafo)
            tribunales = extraer_organos_judiciales(parrafo)

            tribunal_origen = ""
            tribunal_destino = ""
            prov_loc_origen = ""
            prov_loc_destino = ""

            if tribunales:
                provincias_explicitas = extraer_provincias_explicitas(parrafo, tribunales)

                # Separar tipo y localidad del tribunal de origen
                tipo_orig, localidad_orig = separar_tipo_y_localidad(tribunales[0])
                tribunal_origen = tipo_orig

                if 0 in provincias_explicitas:
                    prov_loc_origen = _resolver_localidad(provincias_explicitas[0], tribunales[0])
                else:
                    prov_loc_origen = _resolver_localidad(localidad_orig, tribunales[0])

                if len(tribunales) > 1:
                    # Separar tipo y localidad del tribunal de destino
                    tipo_dest, localidad_dest = separar_tipo_y_localidad(tribunales[1])
                    tribunal_destino = tipo_dest

                    if 1 in provincias_explicitas:
                        prov_loc_destino = _resolver_localidad(provincias_explicitas[1], tribunales[1])
                    else:
                        prov_loc_destino = _resolver_localidad(localidad_dest, tribunales[1])

            filas.append({
                "Participante": participante.replace("Doña ", "").replace("Don ", ""),
                "Tipo de funcionaria": cargo,
                "Tipo de tribunal de origen": tribunal_origen,
                "Tipo de tribunal de destino": tribunal_destino,
                "Provincia/Localidad de origen": prov_loc_origen,
                "Provincia/Localidad de destino": prov_loc_destino,
            })
        except Exception:
            # Saltar párrafos con formato inesperado
            continue

    return filas


def generar_excel_bytes(filas: list[dict]) -> bytes:
    """Genera un archivo Excel en memoria y devuelve los bytes."""
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Participantes"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(vertical="center", wrap_text=True)

    for row_idx, fila in enumerate(filas, start=2):
        for col_idx, col_name in enumerate(COLUMNAS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=fila[col_name])
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    anchos = [30, 25, 40, 40, 35, 35]
    for col_idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = ancho

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def main():
    filas = procesar_pdf("data/PDF2.pdf")

    output_path = "data/participantes.xlsx"
    excel_bytes = generar_excel_bytes(filas)
    with open(output_path, "wb") as f:
        f.write(excel_bytes)

    print(f"✅ Archivo Excel guardado en: {output_path}")
    print(f"   Total de participantes: {len(filas)}")


if __name__ == "__main__":
    main()