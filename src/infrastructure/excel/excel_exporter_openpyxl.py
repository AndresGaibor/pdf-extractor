from io import BytesIO
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from src.domain.models.extracted_data import ExtractedRow
from src.domain.interfaces.extraction_interfaces import IExcelExporter

class OpenPyXLExporter(IExcelExporter):
    COLUMNAS = [
        "Participante",
        "Tipo de funcionaria",
        "Tipo de tribunal de origen",
        "Tipo de tribunal de destino",
        "Provincia/Localidad de origen",
        "Provincia/Localidad de destino",
    ]

    def generate_excel_bytes(self, data: List[ExtractedRow]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Participantes"

        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(self.COLUMNAS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        data_font = Font(name="Calibri", size=11)
        data_alignment = Alignment(vertical="center", wrap_text=True)

        for row_idx, row_obj in enumerate(data, start=2):
            row_dict = row_obj.to_dict()
            for col_idx, col_name in enumerate(self.COLUMNAS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_dict[col_name])
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        anchos = [30, 25, 40, 40, 35, 35]
        for col_idx, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = ancho

        ws.freeze_panes = "A2"

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
